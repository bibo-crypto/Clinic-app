"""Report generation — PDF via ReportLab, Excel via OpenPyXL."""

import os
import json
import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database.db import get_session
from models.invoice import Invoice
from models.appointment import Appointment
from models.patient import Patient

EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


def _get_setting():
    from models.setting import Setting
    session = get_session()
    try:
        s = session.query(Setting).first()
        return s.clinic_name if s else "Clinic"
    finally:
        session.close()


# ─── PDF helpers ──────────────────────────────────────────────────────────────

def _base_pdf(filename: str, title: str):
    path = os.path.join(EXPORT_DIR, filename)
    doc = SimpleDocTemplate(path, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    clinic = _get_setting()
    elements = [
        Paragraph(clinic, styles["Title"]),
        Paragraph(title, styles["Heading2"]),
        Paragraph(str(datetime.date.today()), styles["Normal"]),
        Spacer(1, 0.5 * cm),
    ]
    return doc, elements, path


def _table_style():
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1976d2")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])


def daily_report_pdf(date: datetime.date = None) -> str:
    date = date or datetime.date.today()
    session = get_session()
    try:
        invoices = session.query(Invoice).filter(Invoice.invoice_date == date).all()
        appts = session.query(Appointment).filter(Appointment.date == date).all()
    finally:
        session.close()

    doc, elements, path = _base_pdf(
        f"daily_report_{date}.pdf", f"Daily Report — {date}"
    )

    # Appointments table
    elements.append(Paragraph("Appointments", getSampleStyleSheet()["Heading3"]))
    appt_data = [["#", "Patient", "Doctor", "Time", "Status"]]
    for i, a in enumerate(appts, 1):
        appt_data.append([i, a.patient.full_name if a.patient else "", a.doctor.full_name if a.doctor else "", str(a.time), a.status])
    if len(appt_data) > 1:
        t = Table(appt_data, repeatRows=1)
        t.setStyle(_table_style())
        elements.append(t)
    elements.append(Spacer(1, 0.5 * cm))

    # Revenue summary
    total = sum(inv.total_amount for inv in invoices)
    elements.append(Paragraph(f"Total Revenue: {total:.2f}", getSampleStyleSheet()["Heading3"]))

    doc.build(elements)
    return path


def monthly_report_pdf(year: int, month: int) -> str:
    session = get_session()
    try:
        invoices = session.query(Invoice).all()
        invoices = [
            inv for inv in invoices
            if inv.invoice_date.year == year and inv.invoice_date.month == month
        ]
    finally:
        session.close()

    doc, elements, path = _base_pdf(
        f"monthly_report_{year}_{month:02d}.pdf",
        f"Monthly Report — {year}/{month:02d}"
    )
    data = [["#", "Invoice#", "Patient", "Date", "Total", "Paid"]]
    for i, inv in enumerate(invoices, 1):
        data.append([
            i, inv.id,
            inv.patient.full_name if inv.patient else "",
            str(inv.invoice_date),
            f"{inv.total_amount:.2f}",
            "Yes" if inv.is_paid else "No"
        ])
    if len(data) > 1:
        t = Table(data, repeatRows=1)
        t.setStyle(_table_style())
        elements.append(t)
    total = sum(inv.total_amount for inv in invoices)
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(Paragraph(f"Total: {total:.2f}", getSampleStyleSheet()["Heading3"]))
    doc.build(elements)
    return path


def patient_report_pdf(patient_id: int) -> str:
    session = get_session()
    try:
        patient = session.query(Patient).filter_by(id=patient_id).first()
        records = patient.medical_records if patient else []
        invoices = patient.invoices if patient else []
        pname = patient.full_name if patient else "Unknown"
    finally:
        session.close()

    doc, elements, path = _base_pdf(
        f"patient_report_{patient_id}.pdf",
        f"Patient Report — {pname}"
    )
    # Medical records
    elements.append(Paragraph("Medical Records", getSampleStyleSheet()["Heading3"]))
    data = [["Date", "Complaint", "Diagnosis", "Prescription"]]
    for r in records:
        data.append([str(r.visit_date), r.complaint[:40], r.diagnosis[:40], r.prescription[:40]])
    if len(data) > 1:
        t = Table(data, repeatRows=1)
        t.setStyle(_table_style())
        elements.append(t)
    elements.append(Spacer(1, 0.5 * cm))

    # Billing
    elements.append(Paragraph("Invoices", getSampleStyleSheet()["Heading3"]))
    bdata = [["Date", "Total", "Paid"]]
    for inv in invoices:
        bdata.append([str(inv.invoice_date), f"{inv.total_amount:.2f}", "Yes" if inv.is_paid else "No"])
    if len(bdata) > 1:
        t2 = Table(bdata, repeatRows=1)
        t2.setStyle(_table_style())
        elements.append(t2)
    doc.build(elements)
    return path


# ─── Excel helpers ─────────────────────────────────────────────────────────────

def _header_fill():
    return PatternFill("solid", fgColor="1976D2")


def daily_report_excel(date: datetime.date = None) -> str:
    date = date or datetime.date.today()
    session = get_session()
    try:
        invoices = session.query(Invoice).filter(Invoice.invoice_date == date).all()
    finally:
        session.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Daily {date}"
    headers = ["Invoice#", "Patient", "Date", "Consultation Fee", "Total", "Paid"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center")
    for row, inv in enumerate(invoices, 2):
        ws.cell(row, 1, inv.id)
        ws.cell(row, 2, inv.patient.full_name if inv.patient else "")
        ws.cell(row, 3, str(inv.invoice_date))
        ws.cell(row, 4, inv.consultation_fee)
        ws.cell(row, 5, inv.total_amount)
        ws.cell(row, 6, "Yes" if inv.is_paid else "No")
    path = os.path.join(EXPORT_DIR, f"daily_report_{date}.xlsx")
    wb.save(path)
    return path


def monthly_report_excel(year: int, month: int) -> str:
    session = get_session()
    try:
        all_inv = session.query(Invoice).all()
        invoices = [i for i in all_inv if i.invoice_date.year == year and i.invoice_date.month == month]
    finally:
        session.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"
    headers = ["Invoice#", "Patient", "Date", "Total", "Paid"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill()
    for row, inv in enumerate(invoices, 2):
        ws.cell(row, 1, inv.id)
        ws.cell(row, 2, inv.patient.full_name if inv.patient else "")
        ws.cell(row, 3, str(inv.invoice_date))
        ws.cell(row, 4, inv.total_amount)
        ws.cell(row, 5, "Yes" if inv.is_paid else "No")
    path = os.path.join(EXPORT_DIR, f"monthly_report_{year}_{month:02d}.xlsx")
    wb.save(path)
    return path
